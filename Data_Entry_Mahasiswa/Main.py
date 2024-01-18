import tkinter as tk
from tkinter import ttk, messagebox
import os #import os untuk mengetahui file di device kita
import openpyxl #import openpyxl untuk library python ke xl
from openpyxl.styles import Font

data = []  # data asli awalnya kosong

def data_Add():
    nama = entry_nama.get() #mengambil data string dari entry nama
    programStudi = entry_ProgramStudi.get() #mengambil data string dari combobox program studi
    kelas = entry_kelas.get() #mengambil data string dari combobox kelas
    nim = entry_NIM.get() #mengambil data string dari entry NIM
    jenisKelamin = var.get() #mengambil data string dari radio button jenis kelamin

    if entry_nama.get(): #jika entry nama tidak kosong
        dataTemp = [ #data sementara
            nama,
            programStudi,
            kelas,
            nim,
            jenisKelamin,
        ]  # menyimpan data dari entry di atas kedalam data sementara
        data.append(dataTemp) #menambahkan data sementara ke penampung data sementara

        #menghapus entry saat di enter data
        entry_nama.delete(0, tk.END)
        entry_ProgramStudi.delete(0, tk.END)
        entry_kelas.delete(0, tk.END)
        entry_NIM.delete(0, tk.END)
        var.set(" ")
        #entry_JenisKelamin.delete(0, tk.END)

        #mencetak data data ke dalam tabel user valudation
        for row, item in enumerate(data, start=2):
            for col, value in enumerate(item):
                # membuat tk.Label
                label_data = tk.Label(frame_UserValidation, text=value, padx=10, pady=5)

                # menentukan letak dari isi dari list (nama, nim, fakultas)
                label_data.grid(row=row, column=col)
    
        #membuat file excel dengan alamat peletakannya
        filepath = "C:\\Users\\ASUS\\latihan python\\DataEntry\\Database.xlsx"

        if not os.path.exists(filepath): #jika file belum ada, akan ada pembuatan header pada file excel tersebut
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            heading = ["NAMA LENGKAP", "PROGRAM STUDI", "KELAS", "NIM", "JENIS KELAMIN"]
            bold_font = Font(bold=True) #font
            for col, title in enumerate(heading, start=1):
                sheet.cell(row=1, column=col, value=title).font = bold_font
            # sheet.append(heading)
            workbook.save(filepath)
        
        #mengisi sheet data data ke file excel yang dituju
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        sheet.append([nama, programStudi, kelas, nim, jenisKelamin]) #<<<
        workbook.save(filepath)
    
    else:
        #jika entry kosong sama sekali maka akan muncul alert
        messagebox.showinfo("ERROR!!!", "DATA HARUS ISI MINIMAL NAMA ANDA!")

"""Frame Information"""
root = tk.Tk()
root.title("Data Entry Mahasiswa")
width = 826
height = 500
root.geometry(f"{width}x{height}")

"""Frame Utama"""
frame_UserInformation = tk.LabelFrame(root, text="User Information")
frame_UserInformation.grid(row=0, column=0, padx=10, pady=10)

"""Entry Nama"""
label_nama = tk.Label(frame_UserInformation, text="Nama Lengkap")
entry_nama = tk.Entry(frame_UserInformation)
label_nama.grid(row=0, column=0)
entry_nama.grid(row=1, column=0)

"""Entry Program Studi"""
label_ProgramStudi = tk.Label(frame_UserInformation, text="Program Studi")
entry_ProgramStudi = ttk.Combobox(frame_UserInformation, values=[""," Teknik Informatika ", " Teknik Mesin ", " Teknik Elektro ", " Sistem Informasi "])
label_ProgramStudi.grid(row=0, column=1)
entry_ProgramStudi.grid(row=1, column=1)

"""Entry Kelas"""
label_kelas = tk.Label(frame_UserInformation, text="Kelas")
entry_kelas = ttk.Combobox(frame_UserInformation, values=[""," 03-TPLE001 ", " 03-TPLE002 ", " 03-TPLE003 ", " 03-TPLE004 "])
label_kelas.grid(row=0, column=2)
entry_kelas.grid(row=1, column=2)

"""Entry NIM"""
label_NIM = tk.Label(frame_UserInformation, text="NIM")
entry_NIM = tk.Entry(frame_UserInformation)
label_NIM.grid(row=2, column=0)
entry_NIM.grid(row=3, column=0)

"""Entry Jenis Kelamin"""
label_JenisKelamin = tk.Label(frame_UserInformation, text="Jenis Kelamin")
label_JenisKelamin.grid(row=2, column=1)
frame_jenisKelamlin = tk.LabelFrame(frame_UserInformation)
frame_jenisKelamlin.grid(row=3, column=1)

var = tk.StringVar(value=" ") #menemtukan nilai awal radio button
radio_button1 = tk.Radiobutton(frame_jenisKelamlin, text="Lelaki", variable=var, value="Lelaki")
radio_button1.grid(row=0, column=0)
radio_button2 = tk.Radiobutton(frame_jenisKelamlin, text="Perempuan", variable=var, value="Perempuan")
radio_button2.grid(row=0, column=1)
# entry_JenisKelamin = ttk.Combobox(frame_UserInformation, values=["", " Lelaki ", " Perempuan "])
# entry_JenisKelamin.grid(row=3, column=1)

#membuat padding dan lebar widget yang ada pada frame_UserInformation
for widget in frame_UserInformation.winfo_children():
    widget.grid_configure(padx=20, pady=5)
    widget.config(width=30)
    
#jarak
label_kosong = tk.Label(frame_UserInformation, text=" ")
label_kosong.grid(row=4, column=0)

"""Button Enter"""
buttonEnter = tk.Button(root, text="Enter Data", command= lambda:data_Add())
buttonEnter.grid(row=1, column=0, sticky="news", padx=20, pady=10)

"""Frame Validation"""
frame_UserValidation = tk.LabelFrame(root, text="User Validation", height=300)
frame_UserValidation.grid(row=2, column=0, padx=20, pady=20)

#membuat header pada table data
judul = ["Nama", "Program Studi", "Kelas", "NIM", "Jenis Kelamin"]

for col, isi in enumerate(judul):
    label_judul = tk.Label(frame_UserValidation, text=isi, padx=10)
    label_judul.grid(row=0, column=col)  # menentukan letak judul table
    garis = tk.Label(frame_UserValidation, text="--------------------")
    garis.grid(row=1, column=col)

for widget2 in frame_UserValidation.winfo_children():
    widget2.config(width=19)

root.mainloop()